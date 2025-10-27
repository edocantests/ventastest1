import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO

st.set_page_config(
    page_title="Formato Condicional de Ventas",
    page_icon="📊",
    layout="centered"
)

st.title("📊 Formato Condicional para Ventas")
st.markdown("""
Esta aplicación aplica formato condicional a tu archivo Excel:
- 🟢 **Verde**: Los 5 valores más altos de ventas
- 🔴 **Rojo**: Los 5 valores más bajos de ventas
""")

st.divider()

uploaded_file = st.file_uploader(
    "Sube tu archivo Excel",
    type=['xlsx', 'xls'],
    help="El archivo debe contener una columna llamada 'ventas'"
)

if uploaded_file is not None:
    try:
        df = pd.read_excel(uploaded_file)
        
        if 'ventas' not in df.columns:
            st.error("❌ El archivo no contiene una columna llamada 'ventas'")
            st.info("Columnas disponibles: " + ", ".join(df.columns))
        else:
            st.success("✅ Archivo cargado correctamente")
            st.subheader("Vista previa de los datos")
            st.dataframe(df, use_container_width=True)
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total de filas", len(df))
            with col2:
                st.metric("Venta máxima", f"${df['ventas'].max():,.2f}")
            with col3:
                st.metric("Venta mínima", f"${df['ventas'].min():,.2f}")
            
            st.divider()
            
            if st.button("🎨 Aplicar Formato Condicional", type="primary", use_container_width=True):
                with st.spinner("Aplicando formato condicional..."):
                    output = BytesIO()
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Sheet1')
                    
                    output.seek(0)
                    wb = load_workbook(output)
                    ws = wb.active
                    
                    headers = [cell.value for cell in ws[1]]
                    ventas_col_idx = headers.index('ventas') + 1
                    
                    ventas_values = df['ventas'].dropna()
                    
                    top_5 = ventas_values.nlargest(5).values
                    bottom_5 = ventas_values.nsmallest(5).values
                    
                    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                    
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=ventas_col_idx)
                        cell_value = cell.value
                        
                        if cell_value is not None:
                            if cell_value in top_5:
                                cell.fill = green_fill
                            elif cell_value in bottom_5:
                                cell.fill = red_fill
                    
                    output_final = BytesIO()
                    wb.save(output_final)
                    output_final.seek(0)
                    
                    st.success("✅ Formato aplicado correctamente")
                    
                    st.subheader("📈 Resumen del formato aplicado")
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown("**🟢 Top 5 Ventas (Verde)**")
                        for i, val in enumerate(sorted(top_5, reverse=True), 1):
                            st.write(f"{i}. ${val:,.2f}")
                    
                    with col2:
                        st.markdown("**🔴 Bottom 5 Ventas (Rojo)**")
                        for i, val in enumerate(sorted(bottom_5), 1):
                            st.write(f"{i}. ${val:,.2f}")
                    
                    st.divider()
                    
                    st.download_button(
                        label="⬇️ Descargar archivo con formato",
                        data=output_final,
                        file_name="ventas_formato_condicional.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
    
    except Exception as e:
        st.error(f"❌ Error al procesar el archivo: {str(e)}")
        st.info("Por favor, verifica que el archivo sea un Excel válido.")

else:
    st.info("👆 Sube un archivo Excel para comenzar")
    
    with st.expander("ℹ️ Requisitos del archivo"):
        st.markdown("""
        - El archivo debe ser formato Excel (.xlsx o .xls)
        - Debe contener una columna llamada exactamente **'ventas'**
        - Los valores de ventas deben ser numéricos
        - Se recomienda tener al menos 10 filas de datos
        """)
    
    with st.expander("📝 Ejemplo de estructura"):
        ejemplo_df = pd.DataFrame({
            'producto': ['A', 'B', 'C', 'D', 'E'],
            'ventas': [1000, 2500, 800, 3200, 1500]
        })
        st.dataframe(ejemplo_df, use_container_width=True)

st.divider()
st.caption("Desarrollado con Streamlit 🎈")
